#!/usr/bin/env python3
from argparse import ArgumentParser
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from requests.exceptions import HTTPError
from wallbox import Wallbox

if __name__ == '__main__':
	# Parse arguments
	parser = ArgumentParser('wallbox-report-exporter')
	parser.add_argument('-u', '--user', help='Your myWallbox.com username', required=True)
	parser.add_argument('-p', '--password', help='Your myWallbox.com password', required=True)
	parser.add_argument('-c', '--charger-id', help='Charger ID to generate the report for; required if you have multiple chargers associated', type=int)
	parser.add_argument('-m', '--month', help='Month for the report; defaults to the current month if not specified', type=int)
	parser.add_argument('-y', '--year', help='Year for the report; defaults to the current year if not specified', type=int)
	parser.add_argument('--full-year', help='Generate a report for the entire year instead of a single month', action='store_true', default=False)
	parser.add_argument('-o', '--output', help='Output file name; defaults to YYYY-MM.xlsx or YYYY.xlsx if not specified')
	parser.add_argument('-s', '--summary', help='Include a summary at the end of the table with stats like total energy charged etc.', action='store_true', default=False)
	parser.add_argument('-it', '--italian', help='Use Italian for descriptions and table headings in the report', action='store_true', default=False)
	args = parser.parse_args()

	# Authenticate
	try:
		wallbox = Wallbox(args.user, args.password)
		wallbox.authenticate()
	except HTTPError:
		print('Authentication failed, please check your username and password.')
		exit(1)
	
	# Check if multiple chargers are associated with the account
	charger = wallbox.getChargersList()
	# Specified charger ID
	if args.charger_id is not None:
		if args.charger_id in charger:
			charger = args.charger_id
		else:
			print('The specified charger ID does not exist in your account.')
			exit(2)
	# Only one charger associated with the account
	if isinstance(charger, list) and len(charger) == 1:
		charger = charger[0]
	# Multiple chargers associated and no ID specified
	if isinstance(charger, list):
		print("Multiple chargers found in your account. Please specify the ID of the charger you want to generate the report for.")
		print("Check the README file for instructions on how to find your charger's ID.")
		exit(2)

	# Set start and end dates for the report
	args.month = args.month if args.month is not None else datetime.now().month
	if args.month < 1 or args.month > 12:
		print('Invalid month specified. Please provide a month between 1 and 12.')
		exit(3)
	args.year = args.year if args.year is not None else datetime.now().year
	if args.full_year:
		start = datetime(args.year , 1, 1)
		end = datetime(args.year + 1, 1, 1) - timedelta(seconds=1)
	else:
		start = datetime(args.year, args.month, 1)
		end = datetime(args.year, args.month + 1, 1) - timedelta(seconds=1)

	# Get sessions
	sessions = wallbox.getSessionList(charger, start, end)['data']
	sessions = [session['attributes'] for session in sessions]
	sessions = sorted(sessions, key=lambda x: x['start'])

	# Create worksheet and styling properties
	workbook = Workbook()
	worksheet = workbook.active
	thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	header_font = Font(bold=True)
	header_fill = PatternFill(start_color='f1c232', end_color='f1c232', fill_type='solid')
	data_fill = PatternFill(start_color='a2c4c9', end_color='a2c4c9', fill_type='solid')

	# Insert data
	if args.italian:
		worksheet.append(['Inizio sessione', 'Fine sessione', 'Durata', 'Energia complessiva', 'Energia di rete', 'Energia fotovoltaico', 'Costo', 'Risparmio da fotovoltaico'])
	else:
		worksheet.append(['Session start', 'Session end', 'Duration', 'Total energy', 'Grid energy', 'Solar energy', 'Cost', 'Savings from solar'])
	for session in sessions:
		worksheet.append([
			datetime.fromtimestamp(session['start']).strftime('%d/%m/%Y %H:%M:%S'),
			datetime.fromtimestamp(session['end']).strftime('%d/%m/%Y %H:%M:%S'),
			str(timedelta(seconds=session['time'])),
			f'{round(session['energy'], 2)} {session['energy_unit']}',
			f'{round(session['energy'] - session['green_energy'], 2)} {session['energy_unit']}',
			f'{round(session['green_energy'], 2)} {session['energy_unit']}',
			f'{session['cost_unit']}{round(session['cost'], 2)}',
			f'{session['cost_unit']}{round(session['cost_savings'], 2)}'
		])

	# Apply styling
	for row in worksheet[worksheet.calculate_dimension()]:
		for cell in row:
			cell.border = thin_border
			cell.fill = data_fill
	# Replace styling for header
	for header_cell in worksheet[1]:
		header_cell.font = header_font
		header_cell.fill = header_fill

	# Insert summary
	if args.summary:
		if args.italian:
			summary_header = ['Energia totale', 'Energia di rete totale', 'Energia fotovoltaico totale', 'Costo totale', 'Risparmio totale da fotovoltaico']
		else:
			summary_header = ['Total energy', 'Total grid energy', 'Total solar energy', 'Total cost', 'Total savings from solar']
		summary_data = [
			f"{sum(round(session['energy'], 2) for session in sessions)} {sessions[0]['energy_unit']}",
			f"{sum(round(session['energy'] - session['green_energy'], 2) for session in sessions)} {sessions[0]['energy_unit']}",
			f"{sum(round(session['green_energy'], 2) for session in sessions)} {sessions[0]['energy_unit']}",
			f"{sessions[0]['cost_unit']}{round(sum(session['cost'] for session in sessions), 2)}",
			f"{sessions[0]['cost_unit']}{round(sum(session['cost_savings'] for session in sessions), 2)}"
		]
		end_row = worksheet.max_row
		for i in range(0, len(summary_header)):
			header_cell = worksheet[end_row + i + 3][0]
			data_cell = worksheet[end_row + i + 3][1]
			header_cell.value = summary_header[i]
			header_cell.border = thin_border
			header_cell.font = header_font
			header_cell.fill = header_fill
			data_cell.value = summary_data[i]
			data_cell.border = thin_border
			data_cell.fill = data_fill

	# Set column widths
	worksheet.column_dimensions['B'].width = 22
	worksheet.column_dimensions['C'].width = 22
	worksheet.column_dimensions['D'].width = 15
	worksheet.column_dimensions['E'].width = 15
	worksheet.column_dimensions['F'].width = 15
	worksheet.column_dimensions['G'].width = 15
	worksheet.column_dimensions['H'].width = 15
	worksheet.column_dimensions['I'].width = 17

	worksheet.insert_rows(1)
	worksheet.insert_cols(1)

	# Save
	if args.output is None:
		if args.full_year:
			args.output = f'{args.year}.xlsx'
		else:
			args.output = f'{args.year}-{args.month:02d}.xlsx'
	workbook.save(args.output)